from flask import Blueprint, render_template, request, flash, redirect, url_for, send_file
from flask_login import login_required, current_user
from app.models import Shop, ButtonPress
from app import db
from sqlalchemy import func
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

admin = Blueprint('admin', __name__, url_prefix='/admin')


@admin.route('/')
@login_required
def dashboard():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('main.player'))

    stats_dict = get_formatted_stats()
    shops = Shop.query.filter(Shop.shop_name != 'admin').all()

    return render_template('admin.html',
                           stats=stats_dict,
                           shops=shops,
                           datetime=datetime)


def get_formatted_stats():
    try:
        stats_query = db.session.query(
            Shop.shop_name,
            ButtonPress.button_number,
            func.count(ButtonPress.id).label('play_count')
        ).join(
            ButtonPress, Shop.id == ButtonPress.shop_id, isouter=True
        ).group_by(
            Shop.shop_name,
            ButtonPress.button_number
        ).all()

        # Format statistics
        stats_dict = {}
        video_names = {
            1: '55 inch UHD',
            2: '55 Inch Mini LED',
            3: '65 Inch Gaming QLED',
            4: '65 Inch Mini LED'
        }

        # Initialize all shops with zero counts
        shops = Shop.query.all()
        for shop in shops:
            stats_dict[shop.shop_name] = {
                '55 inch UHD': 0,
                '55 Inch Mini LED': 0,
                '65 Inch Gaming QLED': 0,
                '65 Inch Mini LED': 0
            }

        # Fill in actual counts
        for shop_name, button_number, count in stats_query:
            if button_number in video_names:
                stats_dict[shop_name][video_names[button_number]] = count

        return stats_dict
    except Exception as e:
        print(f"Error getting stats: {str(e)}")
        return {}


@admin.route('/download')
@login_required
def download():
    if not current_user.is_admin:
        flash('Access denied.', 'error')
        return redirect(url_for('main.player'))

    try:
        # Get statistics
        stats_dict = get_formatted_stats()

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Video Statistics"

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='E2EFD9', end_color='E2EFD9', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title
        ws['A1'] = 'Video Player Statistics'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = center_align

        # Add timestamp
        ws['A2'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws.merge_cells('A2:E2')
        ws['A2'].alignment = Alignment(horizontal='left')

        # Write headers
        headers = ['Shop Name', '55 inch UHD', '55 Inch Mini LED', '65 Inch Gaming QLED', '65 Inch Mini LED']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Write data
        row = 4
        for shop_name, counts in stats_dict.items():
            ws.cell(row=row, column=1, value=shop_name).border = border
            ws.cell(row=row, column=1).alignment = center_align

            col = 2
            for video_name in headers[1:]:
                cell = ws.cell(row=row, column=col, value=counts[video_name])
                cell.border = border
                cell.alignment = center_align
                col += 1
            row += 1

        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate timestamp for filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'video_statistics_{timestamp}.xlsx'
        )

    except Exception as e:
        print(f"Export error: {str(e)}")  # For debugging
        flash('Error exporting data to Excel.', 'error')
        return redirect(url_for('admin.dashboard'))


def get_formatted_stats():
    try:
        stats_query = db.session.query(
            Shop.shop_name,
            ButtonPress.button_number,
            func.count(ButtonPress.id).label('play_count')
        ).join(
            ButtonPress, Shop.id == ButtonPress.shop_id, isouter=True
        ).group_by(
            Shop.shop_name,
            ButtonPress.button_number
        ).all()

        stats_dict = {}
        video_names = {
            1: '55 inch UHD',
            2: '55 Inch Mini LED',
            3: '65 Inch Gaming QLED',
            4: '65 Inch Mini LED'
        }

        # Initialize all shops with zero counts
        shops = Shop.query.all()
        for shop in shops:
            stats_dict[shop.shop_name] = {
                '55 inch UHD': 0,
                '55 Inch Mini LED': 0,
                '65 Inch Gaming QLED': 0,
                '65 Inch Mini LED': 0
            }

        # Fill in actual counts
        for shop_name, button_number, count in stats_query:
            if button_number in video_names:
                stats_dict[shop_name][video_names[button_number]] = count

        return stats_dict
    except Exception as e:
        print(f"Error getting stats: {str(e)}")
        return {}

@admin.route('/register', methods=['POST'])
@login_required
def register():
    if not current_user.is_admin:
        flash('Access denied.', 'error')
        return redirect(url_for('main.player'))

    shop_name = request.form.get('shop_name')
    password = request.form.get('password')

    if not shop_name or not password:
        flash('Shop name and password are required.', 'error')
        return redirect(url_for('admin.dashboard'))

    if Shop.query.filter_by(shop_name=shop_name).first():
        flash('Shop name already exists.', 'error')
        return redirect(url_for('admin.dashboard'))

    try:
        new_shop = Shop(shop_name=shop_name)
        new_shop.set_password(password)
        db.session.add(new_shop)
        db.session.commit()
        flash(f'Shop "{shop_name}" registered successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while registering the shop.', 'error')

    return redirect(url_for('admin.dashboard'))


@admin.route('/delete/<int:shop_id>', methods=['POST'])
@login_required
def delete(shop_id):
    if not current_user.is_admin:
        flash('Access denied.', 'error')
        return redirect(url_for('main.player'))

    shop = Shop.query.get_or_404(shop_id)
    if shop.shop_name == 'admin':
        flash('Cannot delete admin account.', 'error')
        return redirect(url_for('admin.dashboard'))

    try:
        ButtonPress.query.filter_by(shop_id=shop.id).delete()
        db.session.delete(shop)
        db.session.commit()
        flash(f'Shop "{shop.shop_name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while deleting the shop.', 'error')

    return redirect(url_for('admin.dashboard'))
